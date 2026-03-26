#Sollicitatielog

import imaplib
import email
from email.header import decode_header
from pathlib import Path
import os
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook

# Instelling laden via dotenv, zodat het paswoord niet openbaar wordt
load_dotenv()
# Zeker pip install python niet vergeten

# .env data inladen
IMAP_SERVER = os.getenv("IMAP_SERVER")
MAIL_ADRES    = os.getenv("MAIL_ADRES")
MAIL_WACHTWOORD = os.getenv("MAIL_WACHTWOORD")
IMAP_MAP       = os.getenv("IMAP_MAP")
DATA_MAP        = os.getenv("DATA_MAP")
XLS_PAD = Path(DATA_MAP) / "sollicitatielog.xlsx"

# foutopvang ontbrekende .env data
ontbrekend = [naam for naam, waarde in {
    "IMAP_SERVER":     IMAP_SERVER,
    "MAIL_ADRES":      MAIL_ADRES,
    "MAIL_WACHTWOORD": MAIL_WACHTWOORD,
    "IMAP_MAP":        IMAP_MAP,
    "DATA_MAP": DATA_MAP,
}.items() if not waarde]

if ontbrekend:
    print(f"❌  Ontbrekende instellingen in .env: {', '.join(ontbrekend)}")
    exit(1)

# mailheader decoderen
def decodeer_header(waarde: str | None) -> str:
    if not waarde:
        return ""
    delen = decode_header(waarde)
    resultaat = []
    for deel, codering in delen:
        if isinstance(deel, bytes):
            resultaat.append(deel.decode(codering or "utf-8", errors="replace"))
        else:
            resultaat.append(deel)
    return " ".join(resultaat)

# mailheader inlezen
def haal_mails_op() -> list[dict]:
    print(f"Verbinding maken met {IMAP_SERVER}...")
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(MAIL_ADRES, MAIL_WACHTWOORD)

    status, _ = mail.select(f'"{IMAP_MAP}"')
    if status != "OK":
        print(f"❌ Map '{IMAP_MAP}' niet gevonden. Controleer de mapnaam.")
        mail.logout()
        return []

    status, berichten = mail.search(None, "ALL")
    if status != "OK" or not berichten[0]:
        print("Geen mails gevonden in deze map.")
        mail.logout()
        return []

    mail_ids = berichten[0].split()
    print(f"✅ {len(mail_ids)} mail(s) gevonden.")

    resultaten = []
    for mail_id in mail_ids:
        try:
            status, data = mail.fetch(mail_id, "(RFC822)")
            if status != "OK":
                print(f"⚠️  Mail {mail_id} kon niet opgehaald worden — overgeslagen.")
                continue

            bericht = email.message_from_bytes(data[0][1])
            onderwerp = decodeer_header(bericht.get("Subject", "(geen onderwerp)"))
            van       = decodeer_header(bericht.get("From", "(onbekend)"))
            van       = email.utils.parseaddr(van)[1]          
            domein    = van.split("@")[1] if "@" in van else ""  
            bedrijf   = domein.split(".")[-2].capitalize() if domein else "(onbekend)"  
            datum_raw = bericht.get("Date", "")

            try:
                datum_obj = email.utils.parsedate_to_datetime(datum_raw)
                datum     = datum_obj.strftime("%d-%m-%Y")
            except Exception:
                datum = datum_raw
            print(f"MAIL OK: {datum} | {van} | {onderwerp}")
            resultaten.append({
                "Datum":     datum,
                "Van":       van,
                "Bedrijf":   bedrijf,
                "Onderwerp": onderwerp,
                
            })
        except Exception as fout:
            print(f"⚠️  Fout bij mail {mail_id}: {type(fout).__name__}: {fout}")
            continue

    mail.logout()
    return resultaten

def sla_op_als_excel(mails: list[dict]) -> None:
    XLS_PAD.parent.mkdir(parents=True, exist_ok=True)

# Eerste keer opstarten: headers alvast klaarzetten zodat openpyxl structuur kent
    if not XLS_PAD.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["Datum", "Van", "Bedrijf", "Onderwerp"])
        wb.save(XLS_PAD)
        
# Opvangen error openstaande excel
    try:
        wb = load_workbook(XLS_PAD)
        ws = wb.active
    
    except PermissionError:
        print("❌ Het Excel-bestand staat open. Sluit het eerst en probeer opnieuw.")
        return
    
# Set is sneller dan lijst voor opzoeken — bij grote logs merk je het verschil
    bestaande = set()
    for rij in ws.iter_rows(min_row=2, values_only=True):
        bestaande.add((rij[0], rij[3]))  # Datum + Onderwerp

    nieuw = [m for m in mails if (m["Datum"], m["Onderwerp"]) not in bestaande]

    if not nieuw:
        print("ℹ️  Geen nieuwe mails om toe te voegen.")
        return

    for m in nieuw:
        ws.append([m["Datum"], m["Van"], m["Bedrijf"], m["Onderwerp"]])

    wb.save(XLS_PAD)
    print(f"✅ {len(nieuw)} nieuwe sollicitatie(s) toegevoegd.")
    print(f"📄 Bestand: {XLS_PAD}")

#hoofdprogramma
if __name__ == "__main__":
    mails = haal_mails_op()
    if mails:
        sla_op_als_excel(mails)
            
    print("\nKlaar! Druk op Enter om af te sluiten.")
    input()